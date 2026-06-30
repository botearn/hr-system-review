"""Excel export helpers (xlsx via openpyxl)."""

from __future__ import annotations

from collections.abc import Iterable
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models.candidate import Candidate
from app.services.matcher import MatchResult

_HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


def _write_header(ws, headers: list[str]) -> None:
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="center")
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def _auto_width(ws, col_widths: list[int]) -> None:
    for idx, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w


def _workbook_to_bytes(wb: Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Candidates
# ---------------------------------------------------------------------------

_CANDIDATE_HEADERS = [
    "ID",
    "姓名",
    "手机",
    "邮箱",
    "城市",
    "行业",
    "工作年限",
    "学历",
    "求职状态",
    "期望薪资下限(k)",
    "期望薪资上限(k)",
    "技能",
    "简历质量分",
    "来源",
    "创建时间",
]

_CANDIDATE_WIDTHS = [6, 10, 14, 24, 10, 14, 10, 8, 10, 14, 14, 40, 10, 10, 20]

_JOB_STATUS_LABEL = {
    "active": "积极求职",
    "watching": "观望中",
    "onboarded": "已入职",
}


def export_candidates_xlsx(candidates: Iterable[Candidate]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "候选人"
    _write_header(ws, _CANDIDATE_HEADERS)
    _auto_width(ws, _CANDIDATE_WIDTHS)

    for row_idx, c in enumerate(candidates, start=2):
        ws.cell(row=row_idx, column=1, value=c.id)
        ws.cell(row=row_idx, column=2, value=c.name)
        ws.cell(row=row_idx, column=3, value=c.phone)
        ws.cell(row=row_idx, column=4, value=c.email)
        ws.cell(row=row_idx, column=5, value=c.city)
        ws.cell(row=row_idx, column=6, value=c.industry)
        ws.cell(row=row_idx, column=7, value=c.years_of_experience)
        ws.cell(row=row_idx, column=8, value=c.education_level)
        ws.cell(row=row_idx, column=9, value=_JOB_STATUS_LABEL.get(c.job_status, c.job_status))
        ws.cell(
            row=row_idx,
            column=10,
            value=float(c.expected_salary_min) if c.expected_salary_min else None,
        )
        ws.cell(
            row=row_idx,
            column=11,
            value=float(c.expected_salary_max) if c.expected_salary_max else None,
        )
        ws.cell(row=row_idx, column=12, value=", ".join(c.skills or []))
        ws.cell(
            row=row_idx,
            column=13,
            value=float(c.resume_quality_score) if c.resume_quality_score else None,
        )
        ws.cell(row=row_idx, column=14, value=c.source)
        ws.cell(
            row=row_idx,
            column=15,
            value=c.created_at.strftime("%Y-%m-%d %H:%M") if c.created_at else None,
        )

    return _workbook_to_bytes(wb)


# ---------------------------------------------------------------------------
# Matches
# ---------------------------------------------------------------------------

_DIM_LABEL = {
    "capability": "能力",
    "skill": "技能",
    "salary": "薪资",
    "industry": "行业",
    "education": "学历",
    "resume_quality": "简历质量",
    "city": "城市",
}


def export_matches_xlsx(
    position_title: str,
    position_city: str | None,
    results: list[MatchResult],
    weights: dict[str, float],
) -> bytes:
    wb = Workbook()

    # ---------- Sheet 1: 匹配结果 ----------
    ws = wb.active
    ws.title = "匹配结果"

    headers = ["排名", "候选人ID", "姓名", "综合分"]
    headers.extend(_DIM_LABEL.get(d, d) for d in weights.keys())
    headers += ["匹配点", "差异点"]

    _write_header(ws, headers)

    dim_keys = list(weights.keys())
    col_widths = [6, 10, 10, 10] + [10] * len(dim_keys) + [50, 50]
    _auto_width(ws, col_widths)

    for row_idx, r in enumerate(results, start=2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        ws.cell(row=row_idx, column=2, value=r.candidate_id)
        ws.cell(row=row_idx, column=3, value=r.candidate_name)
        ws.cell(row=row_idx, column=4, value=round(r.score, 1))
        for i, dim in enumerate(dim_keys):
            ws.cell(row=row_idx, column=5 + i, value=round(r.sub_scores.get(dim, 0.0), 1))
        matched = "; ".join(p.get("detail", "") for p in r.matched_points)
        gaps = "; ".join(p.get("detail", "") for p in r.gap_points)
        cell_m = ws.cell(row=row_idx, column=5 + len(dim_keys), value=matched)
        cell_g = ws.cell(row=row_idx, column=6 + len(dim_keys), value=gaps)
        cell_m.alignment = Alignment(wrap_text=True, vertical="top")
        cell_g.alignment = Alignment(wrap_text=True, vertical="top")

    # ---------- Sheet 2: 元信息 ----------
    meta = wb.create_sheet("元信息")
    meta["A1"] = "岗位"
    meta["B1"] = position_title
    meta["A2"] = "城市"
    meta["B2"] = position_city or ""
    meta["A3"] = "导出时间"
    meta["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    meta["A4"] = "候选人总数"
    meta["B4"] = len(results)
    meta["A6"] = "权重配置"
    meta["A6"].font = Font(bold=True)
    for i, (k, v) in enumerate(weights.items(), start=7):
        meta[f"A{i}"] = _DIM_LABEL.get(k, k)
        meta[f"B{i}"] = v
    _auto_width(meta, [16, 30])

    return _workbook_to_bytes(wb)
